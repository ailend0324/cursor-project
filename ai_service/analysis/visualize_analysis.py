import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import os
import json
from datetime import datetime
from collections import defaultdict
import io
import base64
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from matplotlib.font_manager import FontProperties
import matplotlib as mpl
from wordcloud import WordCloud
import jieba
from sklearn.manifold import TSNE
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# 设置项目根目录
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 设置中文字体
try:
    font_path = '/System/Library/Fonts/PingFang.ttc'  # macOS 上的字体
    if os.path.exists(font_path):
        mpl.rcParams['font.family'] = ['PingFang SC']
    else:
        # 尝试其他常见中文字体
        font_paths = [
            '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',  # Linux
            'C:/Windows/Fonts/simhei.ttf',  # Windows
            'C:/Windows/Fonts/simsun.ttc'  # Windows
        ]
        for path in font_paths:
            if os.path.exists(path):
                font = FontProperties(fname=path)
                plt.rcParams['font.family'] = font.get_name()
                break
except:
    print("无法设置中文字体，图表中的中文可能无法正确显示")

def load_data(file_path):
    """加载Excel文件数据"""
    try:
        print(f"Loading data from {file_path}...")
        df = pd.read_excel(file_path)
        print(f"Successfully loaded {len(df)} records.")
        return df
    except Exception as e:
        print(f"Error loading data: {e}")
        return None

def plot_to_image():
    """Convert matplotlib plot to base64 image for HTML embedding"""
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=100)
    buf.seek(0)
    image_base64 = base64.b64encode(buf.getvalue()).decode('utf-8')
    plt.close()
    return image_base64

def visualize_question_categories(df):
    """Visualize question categories"""
    # Filter for user messages only
    user_df = df[df['sender_type'] == 1.0]
    
    # Function to categorize questions
    def categorize_question(text):
        if not isinstance(text, str):
            return "其他类"
        
        categories = {
            "产品咨询类": ["价格", "机型", "型号", "回收", "估价", "报价", "检测", "评估"],
            "服务支持类": ["账号", "登录", "支付", "运费", "物流", "订单", "取消", "退款", "退回"],
            "技术问题类": ["报错", "故障", "操作", "无法", "问题", "失败", "解锁", "恢复"],
            "业务咨询类": ["合作", "商务", "企业", "批量", "团购", "代理"]
        }
        
        for category, keywords in categories.items():
            for keyword in keywords:
                if keyword in text:
                    return category
        
        return "其他类"
    
    # Categorize questions
    user_df['category'] = user_df['send_content'].fillna("").apply(categorize_question)
    category_counts = user_df['category'].value_counts()
    
    # Plotting
    plt.figure(figsize=(10, 6))
    sns.barplot(x=category_counts.index, y=category_counts.values)
    plt.title('Distribution of Question Categories')
    plt.xlabel('Category')
    plt.ylabel('Count')
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    return {
        'image': plot_to_image(),
        'data': category_counts.to_dict()
    }

def visualize_top_questions(df):
    """Visualize top user questions"""
    # Filter for user messages only
    user_df = df[df['sender_type'] == 1.0]
    
    # Count question frequencies
    question_counts = user_df['send_content'].value_counts().head(10)
    
    # Plotting
    plt.figure(figsize=(12, 6))
    sns.barplot(x=question_counts.values, y=question_counts.index)
    plt.title('Top 10 User Questions')
    plt.xlabel('Count')
    plt.ylabel('Question')
    plt.tight_layout()
    
    return {
        'image': plot_to_image(),
        'data': question_counts.to_dict()
    }

def visualize_conversation_duration(df):
    """Visualize conversation duration distribution"""
    # Create a copy of the dataframe with only the rows having valid start and end times
    time_df = df.dropna(subset=['user_start_time', 'user_end_time']).copy()
    
    # Convert to datetime
    time_df['user_start_time'] = pd.to_datetime(time_df['user_start_time'])
    time_df['user_end_time'] = pd.to_datetime(time_df['user_end_time'])
    
    # Calculate duration in minutes
    time_df['duration_minutes'] = (time_df['user_end_time'] - time_df['user_start_time']).dt.total_seconds() / 60
    
    # Filter out negative durations and unreasonably long durations (more than 2 hours)
    time_df = time_df[(time_df['duration_minutes'] >= 0) & (time_df['duration_minutes'] <= 120)]
    
    # Categorize duration into bins
    duration_bins = [0, 1, 3, 5, 10, 15, 30, 60, 120]
    duration_labels = ['<1min', '1-3min', '3-5min', '5-10min', '10-15min', '15-30min', '30-60min', '1-2hrs']
    time_df['duration_category'] = pd.cut(time_df['duration_minutes'], bins=duration_bins, labels=duration_labels)
    
    # Count conversations in each duration category
    duration_counts = time_df['duration_category'].value_counts().sort_index()
    
    # Plotting
    plt.figure(figsize=(12, 6))
    sns.barplot(x=duration_counts.index, y=duration_counts.values)
    plt.title('Conversation Duration Distribution')
    plt.xlabel('Duration')
    plt.ylabel('Count')
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    return {
        'image': plot_to_image(),
        'data': duration_counts.to_dict()
    }

def visualize_human_transfer(df):
    """Visualize human transfer patterns"""
    # Filter for user messages only
    user_df = df[df['sender_type'] == 1.0].copy()
    
    # Define keywords for human transfer requests
    transfer_keywords = ["人工", "转人工", "真人", "客服", "人工服务", "转接", "转接人工", "请转人工"]
    
    # Create a function to detect transfer requests
    def is_transfer_request(text):
        if not isinstance(text, str):
            return False
        return any(keyword in text for keyword in transfer_keywords)
    
    # Mark transfer requests
    user_df['is_transfer_request'] = user_df['send_content'].apply(is_transfer_request)
    
    # Analyze at which turn users request transfer
    transfer_df = user_df[user_df['is_transfer_request']].copy()
    
    if len(transfer_df) > 0:
        # Get the first transfer request in each conversation
        transfer_df = transfer_df.sort_values(['touch_id', 'seq_no'])
        first_transfers = transfer_df.drop_duplicates('touch_id')
        
        # Bin the turns
        turn_bins = [0, 1, 2, 3, 5, 10, 20, 100]
        turn_labels = ['First msg', 'Turn 2', 'Turn 3', 'Turns 4-5', 'Turns 6-10', 'Turns 11-20', '20+ turns']
        first_transfers['turn_category'] = pd.cut(first_transfers['seq_no'], bins=turn_bins, labels=turn_labels)
        
        # Count by turn category
        turn_counts = first_transfers['turn_category'].value_counts().sort_index()
        
        # Plotting
        plt.figure(figsize=(12, 6))
        sns.barplot(x=turn_counts.index, y=turn_counts.values)
        plt.title('Human Transfer Requests by Conversation Turn')
        plt.xlabel('Turn')
        plt.ylabel('Count')
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        return {
            'image': plot_to_image(),
            'data': turn_counts.to_dict()
        }
    else:
        return {
            'image': None,
            'data': {}
        }

def visualize_conversation_turns(df):
    """Visualize distribution of conversation turns"""
    # Group by conversation ID (touch_id)
    conversation_groups = df.groupby('touch_id')
    
    # Calculate the number of dialogue turns per conversation
    dialogue_turns = conversation_groups.size()
    
    # Create bins for dialogue turns
    turn_bins = [0, 3, 5, 10, 15, 20, 30, 50, 100, dialogue_turns.max() + 1]
    turn_labels = ['1-3', '4-5', '6-10', '11-15', '16-20', '21-30', '31-50', '51-100', '100+']
    turn_categories = pd.cut(dialogue_turns, bins=turn_bins, labels=turn_labels)
    
    # Count conversations in each turn category
    turn_counts = turn_categories.value_counts().sort_index()
    
    # Plotting
    plt.figure(figsize=(12, 6))
    sns.barplot(x=turn_counts.index, y=turn_counts.values)
    plt.title('Distribution of Conversation Lengths (Turns)')
    plt.xlabel('Number of Turns')
    plt.ylabel('Number of Conversations')
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    return {
        'image': plot_to_image(),
        'data': turn_counts.to_dict()
    }

def generate_html_report(visualizations):
    """Generate HTML report with visualizations"""
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Chat Analysis Report</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 0;
                padding: 20px;
                color: #333;
            }}
            .container {{
                max-width: 1200px;
                margin: 0 auto;
            }}
            .header {{
                background-color: #4a6fa5;
                color: white;
                padding: 20px;
                text-align: center;
                margin-bottom: 30px;
                border-radius: 5px;
            }}
            .section {{
                margin-bottom: 40px;
                background-color: #f9f9f9;
                padding: 20px;
                border-radius: 5px;
                box-shadow: 0 2px 5px rgba(0,0,0,0.1);
            }}
            .chart {{
                text-align: center;
                margin-bottom: 20px;
            }}
            .chart img {{
                max-width: 100%;
                height: auto;
            }}
            h1, h2, h3 {{
                color: #2c3e50;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 20px;
            }}
            th, td {{
                padding: 10px;
                border: 1px solid #ddd;
                text-align: left;
            }}
            th {{
                background-color: #4a6fa5;
                color: white;
            }}
            tr:nth-child(even) {{
                background-color: #f2f2f2;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Customer Service Conversation Analysis Report</h1>
                <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            
            <div class="section">
                <h2>Question Categories</h2>
                <div class="chart">
                    <img src="data:image/png;base64,{visualizations['question_categories']['image']}" alt="Question Categories">
                </div>
                <table>
                    <tr>
                        <th>Category</th>
                        <th>Count</th>
                    </tr>
    """
    
    for category, count in visualizations['question_categories']['data'].items():
        html_content += f"""
                    <tr>
                        <td>{category}</td>
                        <td>{count}</td>
                    </tr>
        """
    
    html_content += """
                </table>
            </div>
            
            <div class="section">
                <h2>Top 10 User Questions</h2>
                <div class="chart">
                    <img src="data:image/png;base64,{0}" alt="Top User Questions">
                </div>
                <table>
                    <tr>
                        <th>Question</th>
                        <th>Count</th>
                    </tr>
    """.format(visualizations['top_questions']['image'])
    
    for question, count in visualizations['top_questions']['data'].items():
        # Truncate long questions
        display_question = question if len(str(question)) < 50 else str(question)[:47] + "..."
        html_content += f"""
                    <tr>
                        <td>{display_question}</td>
                        <td>{count}</td>
                    </tr>
        """
    
    html_content += """
                </table>
            </div>
            
            <div class="section">
                <h2>Conversation Duration Distribution</h2>
                <div class="chart">
                    <img src="data:image/png;base64,{0}" alt="Conversation Duration">
                </div>
                <table>
                    <tr>
                        <th>Duration</th>
                        <th>Count</th>
                    </tr>
    """.format(visualizations['conversation_duration']['image'])
    
    for duration, count in visualizations['conversation_duration']['data'].items():
        html_content += f"""
                    <tr>
                        <td>{duration}</td>
                        <td>{count}</td>
                    </tr>
        """
    
    html_content += """
                </table>
            </div>
            
            <div class="section">
                <h2>Conversation Turns Distribution</h2>
                <div class="chart">
                    <img src="data:image/png;base64,{0}" alt="Conversation Turns">
                </div>
                <table>
                    <tr>
                        <th>Number of Turns</th>
                        <th>Count</th>
                    </tr>
    """.format(visualizations['conversation_turns']['image'])
    
    for turns, count in visualizations['conversation_turns']['data'].items():
        html_content += f"""
                    <tr>
                        <td>{turns}</td>
                        <td>{count}</td>
                    </tr>
        """
    
    html_content += """
                </table>
            </div>
    """
    
    if visualizations['human_transfer']['image']:
        html_content += """
            <div class="section">
                <h2>Human Transfer Requests</h2>
                <div class="chart">
                    <img src="data:image/png;base64,{0}" alt="Human Transfer Requests">
                </div>
                <table>
                    <tr>
                        <th>Turn</th>
                        <th>Count</th>
                    </tr>
        """.format(visualizations['human_transfer']['image'])
        
        for turn, count in visualizations['human_transfer']['data'].items():
            html_content += f"""
                        <tr>
                            <td>{turn}</td>
                            <td>{count}</td>
                        </tr>
            """
        
        html_content += """
                </table>
            </div>
        """
    
    html_content += """
        </div>
    </body>
    </html>
    """
    
    return html_content

def generate_excel_report(df, visualizations):
    """Generate Excel report with data and charts"""
    wb = Workbook()
    
    # Summary sheet
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    
    # Set header
    summary_sheet['A1'] = "Customer Service Conversation Analysis"
    summary_sheet['A1'].font = Font(size=16, bold=True)
    summary_sheet.merge_cells('A1:H1')
    
    summary_sheet['A2'] = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    summary_sheet['A2'].font = Font(italic=True)
    summary_sheet.merge_cells('A2:H2')
    
    # Basic stats
    summary_sheet['A4'] = "Basic Statistics"
    summary_sheet['A4'].font = Font(size=14, bold=True)
    summary_sheet.merge_cells('A4:H4')
    
    stats = [
        ["Total Records", len(df)],
        ["Total Conversations", df['touch_id'].nunique()],
        ["Unique Users", df['user_name'].nunique()],
        ["Date Range", f"{df['create_time'].min().strftime('%Y-%m-%d')} to {df['create_time'].max().strftime('%Y-%m-%d')}"]
    ]
    
    for i, (label, value) in enumerate(stats, 5):
        summary_sheet[f'A{i}'] = label
        summary_sheet[f'B{i}'] = value
    
    # Question Categories
    question_sheet = wb.create_sheet("Question Categories")
    question_sheet['A1'] = "Question Categories"
    question_sheet['A1'].font = Font(size=16, bold=True)
    question_sheet.merge_cells('A1:C1')
    
    question_sheet['A3'] = "Category"
    question_sheet['B3'] = "Count"
    question_sheet['C3'] = "Percentage"
    
    for col in ['A3', 'B3', 'C3']:
        question_sheet[col].font = Font(bold=True)
    
    total_count = sum(visualizations['question_categories']['data'].values())
    
    for i, (category, count) in enumerate(visualizations['question_categories']['data'].items(), 4):
        question_sheet[f'A{i}'] = category
        question_sheet[f'B{i}'] = count
        question_sheet[f'C{i}'] = f"{count/total_count:.2%}"
    
    # Create chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Question Categories"
    chart.x_axis.title = "Category"
    chart.y_axis.title = "Count"
    
    data = Reference(question_sheet, min_col=2, min_row=3, max_row=3+len(visualizations['question_categories']['data']))
    cats = Reference(question_sheet, min_col=1, min_row=4, max_row=3+len(visualizations['question_categories']['data']))
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    question_sheet.add_chart(chart, "E4")
    
    # Top Questions
    top_questions_sheet = wb.create_sheet("Top Questions")
    top_questions_sheet['A1'] = "Top 10 User Questions"
    top_questions_sheet['A1'].font = Font(size=16, bold=True)
    top_questions_sheet.merge_cells('A1:C1')
    
    top_questions_sheet['A3'] = "Question"
    top_questions_sheet['B3'] = "Count"
    
    for col in ['A3', 'B3']:
        top_questions_sheet[col].font = Font(bold=True)
    
    for i, (question, count) in enumerate(visualizations['top_questions']['data'].items(), 4):
        display_question = question if len(str(question)) < 50 else str(question)[:47] + "..."
        top_questions_sheet[f'A{i}'] = display_question
        top_questions_sheet[f'B{i}'] = count
    
    # Conversation Duration
    duration_sheet = wb.create_sheet("Conversation Duration")
    duration_sheet['A1'] = "Conversation Duration Distribution"
    duration_sheet['A1'].font = Font(size=16, bold=True)
    duration_sheet.merge_cells('A1:C1')
    
    duration_sheet['A3'] = "Duration"
    duration_sheet['B3'] = "Count"
    duration_sheet['C3'] = "Percentage"
    
    for col in ['A3', 'B3', 'C3']:
        duration_sheet[col].font = Font(bold=True)
    
    total_count = sum(visualizations['conversation_duration']['data'].values())
    
    for i, (duration, count) in enumerate(visualizations['conversation_duration']['data'].items(), 4):
        duration_sheet[f'A{i}'] = duration
        duration_sheet[f'B{i}'] = count
        duration_sheet[f'C{i}'] = f"{count/total_count:.2%}"
    
    # Create chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Conversation Duration"
    chart.x_axis.title = "Duration"
    chart.y_axis.title = "Count"
    
    data = Reference(duration_sheet, min_col=2, min_row=3, max_row=3+len(visualizations['conversation_duration']['data']))
    cats = Reference(duration_sheet, min_col=1, min_row=4, max_row=3+len(visualizations['conversation_duration']['data']))
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    duration_sheet.add_chart(chart, "E4")
    
    # Conversation Turns
    turns_sheet = wb.create_sheet("Conversation Turns")
    turns_sheet['A1'] = "Conversation Turns Distribution"
    turns_sheet['A1'].font = Font(size=16, bold=True)
    turns_sheet.merge_cells('A1:C1')
    
    turns_sheet['A3'] = "Number of Turns"
    turns_sheet['B3'] = "Count"
    turns_sheet['C3'] = "Percentage"
    
    for col in ['A3', 'B3', 'C3']:
        turns_sheet[col].font = Font(bold=True)
    
    total_count = sum(visualizations['conversation_turns']['data'].values())
    
    for i, (turns, count) in enumerate(visualizations['conversation_turns']['data'].items(), 4):
        turns_sheet[f'A{i}'] = turns
        turns_sheet[f'B{i}'] = count
        turns_sheet[f'C{i}'] = f"{count/total_count:.2%}"
    
    # Create chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Conversation Turns"
    chart.x_axis.title = "Number of Turns"
    chart.y_axis.title = "Count"
    
    data = Reference(turns_sheet, min_col=2, min_row=3, max_row=3+len(visualizations['conversation_turns']['data']))
    cats = Reference(turns_sheet, min_col=1, min_row=4, max_row=3+len(visualizations['conversation_turns']['data']))
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    turns_sheet.add_chart(chart, "E4")
    
    # Human Transfer
    if visualizations['human_transfer']['data']:
        transfer_sheet = wb.create_sheet("Human Transfer")
        transfer_sheet['A1'] = "Human Transfer Requests"
        transfer_sheet['A1'].font = Font(size=16, bold=True)
        transfer_sheet.merge_cells('A1:C1')
        
        transfer_sheet['A3'] = "Turn"
        transfer_sheet['B3'] = "Count"
        transfer_sheet['C3'] = "Percentage"
        
        for col in ['A3', 'B3', 'C3']:
            transfer_sheet[col].font = Font(bold=True)
        
        total_count = sum(visualizations['human_transfer']['data'].values())
        
        for i, (turn, count) in enumerate(visualizations['human_transfer']['data'].items(), 4):
            transfer_sheet[f'A{i}'] = turn
            transfer_sheet[f'B{i}'] = count
            transfer_sheet[f'C{i}'] = f"{count/total_count:.2%}"
        
        # Create chart
        chart = BarChart()
        chart.type = "col"
        chart.title = "Human Transfer Requests"
        chart.x_axis.title = "Turn"
        chart.y_axis.title = "Count"
        
        data = Reference(transfer_sheet, min_col=2, min_row=3, max_row=3+len(visualizations['human_transfer']['data']))
        cats = Reference(transfer_sheet, min_col=1, min_row=4, max_row=3+len(visualizations['human_transfer']['data']))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        transfer_sheet.add_chart(chart, "E4")
    
    # Adjust column widths
    for sheet in wb:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def main():
    # 文件路径
    file_path = os.path.join(PROJECT_ROOT, "data", "merged_chat_records.xlsx")
    
    # 加载数据
    df = load_data(file_path)
    if df is None:
        return
    
    # Convert time columns to datetime
    if 'create_time' in df.columns:
        df['create_time'] = pd.to_datetime(df['create_time'])
    
    # Create visualizations
    visualizations = {
        'question_categories': visualize_question_categories(df),
        'top_questions': visualize_top_questions(df),
        'conversation_duration': visualize_conversation_duration(df),
        'conversation_turns': visualize_conversation_turns(df),
        'human_transfer': visualize_human_transfer(df)
    }
    
    # Generate HTML report
    html_content = generate_html_report(visualizations)
    html_output_path = os.path.join(PROJECT_ROOT, "analysis", "chat_analysis_report.html")
    with open(html_output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"HTML report saved to {html_output_path}")
    
    # Generate Excel report
    workbook = generate_excel_report(df, visualizations)
    excel_output_path = os.path.join(PROJECT_ROOT, "analysis", "chat_analysis_report.xlsx")
    workbook.save(excel_output_path)
    print(f"Excel report saved to {excel_output_path}")

if __name__ == "__main__":
    main() 